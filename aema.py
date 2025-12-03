def create_gh_issue(token = None, owner = None, repo = None):
    import pandas as pd
    import requests
    import time
    
    all_vocabs = collect_vocab_2()
    list_and_values = pd.concat(all_vocabs, axis=1)
    result_val = [
        f"{col} | {val}"
        for col in list_and_values.columns
        for val in list_and_values[col].dropna()
    ]
    # noms de colonnes
    result_col = [
        f"{col}"
        for col in list_and_values.columns
    ]
    # fusionne + ordonne + supprime les doublons
    joinedlist = result_col + result_val
    joinedlist = sorted(joinedlist)
    # print(type(joinedlist))
    joinedlist = list(set(joinedlist))

    # === SETTINGS ===
    # token = "ghp_your_personal_access_token_here"  # 🔒 replace with your real token
    # owner = "your-github-username"                 # e.g. "iramat"
    # repo = "your-repo-name"                        # e.g. "aema-private"

    # Example list of issue titles
    # issue_titles = [
    #     "Fix data import for communes",
    #     "Add new vocabulary for object types",
    #     "Update documentation on API endpoints"
    # ]
    issue_titles = joinedlist
    # issue_titles = ['matieres-matiere | Or et alliages associés', 'observations-observation | Observation', 'contextes-contexte', 'series-serie', 'observations-observation', 'localisationtypes-type | Fouille', 'depots-lieuDepot | Cabinet des médailles - BNF Richelieu', 'tresors-tresor', 'denominations-denomination | Autre', 'mesuretypes-type | Au', 'emetteurs-emetteurAutorite | M. Valerius Messalla', 'communes-commune', 'mesuretypes-type | Pt/Pd', 'matieres-matiere | Alliages cuivreux | laiton', 'mesuretypes-type | Ag', 'ateliers-atelierEmission | Rome', 'descriptiftypes-type', 'matieres-matiere | Alliages cuivreux | bronze', 'matieres-matiere | Alliages cuivreux | cuivre', 'tresors-tresor | Autre', 'series-serie | Verca', 'depots-lieuDepot | Pépin le Bref', 'collections-collection | Luynes', 'mesuretypes-type | Pt', 'mesuretypes-type | Sb', 'localisationtypes-type', 'matieres-estMetal', 'localisationtypes-type | Découverte ancienne', 'fonctions-fonction | Autre', 'techniques-technique | Monnaie plaquée', 'communes-commune | Chantenay Saint-Imbert', 'mesuretypes-type | Rh', 'techniques-technique', 'matieres-matiere | Argent et alliages associés', 'series-serie | RRC 366/2a', 'descriptiftypes-type | Types de descriptif', 'emetteurs-emetteurAutorite | Marcus Junius Brutus', 'objetsoustypes-sousType', 'periodes-periode', 'matieres-matiere | Autres', 'objetsoustypes-sousType | Celtique', 'objettypes-type', 'matieres-estMetal | True', 'series-serie | Vercingetorix', 'ateliers-atelierEmission | Saint-Denis', 'ateliers-atelierEmission | Saint-Denis', 'ateliers-atelierEmission | Saint-Denis', 'ateliers-atelierEmission | Saint-Denis', 'ateliers-atelierEmission | Saint-Denis', 'ateliers-atelierEmission | Saint-Denis', 'ateliers-atelierEmission | Saint-Denis']


    # print(issue_titles[80])

    # Markdown bodies for each issue
    issue_bodies = [
"""## Listes

Dernier état des listes AeMA: [listes_et_valeurs_latest.tsv](https://github.com/iramat/aema-private/blob/main/listes/listes_et_valeurs_latest.tsv).
"""
    ]

    # GitHub API endpoint
    url = f"https://api.github.com/repos/{owner}/{repo}/issues"

    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json"
    }

    # === LOOP TO CREATE ISSUES ===
    for title in issue_titles:
        # print(title)
        payload = {
            "title": f"LISTE | {title}",
            "body": issue_bodies[0],
            "labels": ["liste"]  # 👈 Add your tag/label here
        }

        response = requests.post(url, json=payload, headers=headers)

        if response.status_code == 201:
            print(f"✅ Created issue: {title}")
        else:
            print(f"❌ Failed to create issue '{title}': {response.status_code} {response.text}")
            
        time.sleep(3)

def collect_vocab_2(
    vocabs=[
        "ateliers", "objettypes", "collections", "communes", "contextes", "denominations",
        "mesuretypes", "emetteurs", "fonctions", "depots", "matieres", "observations",
        "pays", "periodes", "series", "objetsoustypes", "techniques", "tresors",
        "types", "descriptiftypes", "localisationtypes", "zones"
    ],
    root_url='https://aema.huma-num.fr/back/public/api/'
):
    import pandas as pd
    import requests

    dfs = []

    for vocab in vocabs:
        url = f"{root_url}{vocab}"
        try:
            response = requests.get(url)
            response.raise_for_status()
            data = response.json()

            # Remove the 'id' field and rename other keys with the vocab prefix
            processed_data = []
            for item in data:
                item.pop('id', None)  # Remove 'id' safely
                renamed_item = {f"{vocab}-{k}": v for k, v in item.items()}
                processed_data.append(renamed_item)

            # Create a DataFrame from the processed data
            df = pd.DataFrame(processed_data)
            dfs.append(df)

        except Exception as e:
            print(f"Error fetching {vocab}: {e}")
            dfs.append(pd.DataFrame())  # empty placeholder if error

    return dfs

def collect_vocab(vocabs = [
        "ateliers", "objettypes", "collections", "communes", "contextes", "denominations",
        "mesuretypes", "emetteurs", "fonctions", "depots", "matieres", "observations",
        "pays", "periodes", "series", "objetsoustypes", "techniques", "tresors",
        "types", "descriptiftypes", "localisationtypes", "zones"
    ], root_url = 'https://aema.huma-num.fr/back/public/api/'):
# === 2. Fetch data ===
    import pandas as pd
    import requests

    dfs = []
    for vocab in vocabs:
        url = f"{root_url}{vocab}"
        try:
            response = requests.get(url)
            response.raise_for_status()
            data = response.json()
            for item in data:
                item.pop('id', None)
            dfs.append(pd.DataFrame(data))
        except Exception as e:
            print(f"Error fetching {vocab}: {e}")
            dfs.append(pd.DataFrame())
    return(dfs)

def create_bulkupload(output_path="aeamena_data.xlsx", root_url = 'https://aema.huma-num.fr/back/public/api/', vocabs = [
        "ateliers", "objettypes", "collections", "communes", "contextes", "denominations",
        "mesuretypes", "emetteurs", "fonctions", "depots", "matieres", "observations",
        "pays", "periodes", "series", "objetsoustypes", "techniques", "tresors",
        "types", "descriptiftypes", "localisationtypes", "zones"
    ]
):
    """Fetches AeMA vocabs data, writes them to Excel sheets, and adds dropdowns in a summary sheet."""
    # TODO: transpose rows and columns in the Summary sheet

    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation    

    dfs=collect_vocab(vocabs=vocabs, root_url=root_url)

    # # === 2. Fetch data ===
    # dfs = []
    # for vocab in vocabs:
    #     url = f"{root_url}{vocab}"
    #     try:
    #         response = requests.get(url)
    #         response.raise_for_status()
    #         data = response.json()
    #         for item in data:
    #             item.pop('id', None)
    #         dfs.append(pd.DataFrame(data))
    #     except Exception as e:
    #         print(f"Error fetching {vocab}: {e}")
    #         dfs.append(pd.DataFrame())

    # === 3. Create summary sheet ===
    summary_rows = [
        {"sheet": sheet_name, "column": col}
        for sheet_name, df in zip(vocabs, dfs)
        for col in df.columns
    ]
    summary_df = pd.DataFrame(summary_rows)

    # === 4. Write to Excel ===
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        for name, df in zip(vocabs, dfs):
            df.to_excel(writer, sheet_name=name, index=False)

    # === 5. Add dropdowns ===
    wb = load_workbook(output_path)
    summary_ws = wb["Summary"]
    dropdown_col_idx = 3
    summary_ws.cell(row=1, column=dropdown_col_idx, value="Dropdown")

    for i, row in enumerate(summary_ws.iter_rows(min_row=2, max_row=summary_ws.max_row, max_col=2), start=2):
        sheet_name, column_name = row[0].value, row[1].value
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            header = [cell.value for cell in sheet[1]]
            if column_name in header:
                col_idx = header.index(column_name) + 1
                col_data = list(sheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, values_only=True))[0]
                values = list(dict.fromkeys([str(v) for v in col_data if v]))

                # Build a dropdown list without exceeding Excel's formula limit
                joined = ""
                for value in values:
                    if len(joined) + len(value) + 1 <= 254:
                        joined += ("," if joined else "") + value
                    else:
                        break

                if joined:
                    dropdown_formula = f'"{joined}"'
                    dv = DataValidation(type="list", formula1=dropdown_formula, showDropDown=False)
                    summary_ws.add_data_validation(dv)
                    dv.add(summary_ws.cell(row=i, column=dropdown_col_idx))

    # === 6. Move Summary to first sheet and save ===
    wb._sheets.insert(0, wb._sheets.pop(wb.sheetnames.index("Summary")))
    wb.save(output_path)
    return(wb)
    # print(f"Excel file saved to: {output_path}")

def gallica_api(api_root = 'https://gallica.bnf.fr/services/OAIRecord?ark=', ark_id= 'btv1b104536783', verbose = True):
    """
    Collect Gallica data from different APIs

    :param api_root: API root path
    :param ark_id: Ark
    """
    import requests
    import xml.etree.ElementTree as ET
    import json

    GALLICA_URL = api_root + ark_id

    if api_root == 'https://gallica.bnf.fr/services/OAIRecord?ark=':
      if(verbose):
         print("collect parametric data")
      resp = requests.get(GALLICA_URL)
      if resp.status_code == 200:
          xml_data = resp.text
          root = ET.fromstring(xml_data)
          data = {
              "identifier": root.find(".//identifier").text if root.find(".//identifier") is not None else None,
              "title": root.find(".//dc:title", namespaces={"dc": "http://purl.org/dc/elements/1.1/"}).text if root.find(".//dc:title", namespaces={"dc": "http://purl.org/dc/elements/1.1/"}) is not None else None,
              "format": [fmt.text for fmt in root.findall(".//dc:format", namespaces={"dc": "http://purl.org/dc/elements/1.1/"})],
              "language": root.find(".//dc:language", namespaces={"dc": "http://purl.org/dc/elements/1.1/"}).text if root.find(".//dc:language", namespaces={"dc": "http://purl.org/dc/elements/1.1/"}) is not None else None,
              "relation": root.find(".//dc:relation", namespaces={"dc": "http://purl.org/dc/elements/1.1/"}).text if root.find(".//dc:relation", namespaces={"dc": "http://purl.org/dc/elements/1.1/"}) is not None else None,
              "types": [t.text for t in root.findall(".//dc:type", namespaces={"dc": "http://purl.org/dc/elements/1.1/"})],
              "source": root.find(".//dc:source", namespaces={"dc": "http://purl.org/dc/elements/1.1/"}).text if root.find(".//dc:source", namespaces={"dc": "http://purl.org/dc/elements/1.1/"}) is not None else None,
              "rights": [r.text for r in root.findall(".//dc:rights", namespaces={"dc": "http://purl.org/dc/elements/1.1/"})],
              "description": root.find(".//dc:description", namespaces={"dc": "http://purl.org/dc/elements/1.1/"}).text if root.find(".//dc:description", namespaces={"dc": "http://purl.org/dc/elements/1.1/"}) is not None else None,
          }
          json_output = json.dumps(data, indent=4, ensure_ascii=False).encode('utf8')
          return(json_output.decode())
      else:
          print(f"Error {resp.status_code}: Unable to fetch data from Gallica API")

    if api_root == 'https://gallica.bnf.fr/ark:/12148/':
      if(verbose):
         print("collect image thumbnail")
      GALLICA_URL = GALLICA_URL + '/thumbnail'
      resp = requests.get(GALLICA_URL)
      if resp.status_code == 200:
         return(resp.content)
      else:
          print(f"Failed to load image. HTTP Status Code: {resp.status_code}")

